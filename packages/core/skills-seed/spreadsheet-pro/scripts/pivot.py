#!/usr/bin/env python3
"""数据透视表。用法：python pivot.py input.xlsx output.xlsx --rows "部门" --cols "季度" --vals "销售额" --agg sum"""
import sys; from pathlib import Path
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError: print("需要 openpyxl。pip install openpyxl"); sys.exit(1)

def main():
    import argparse
    p = argparse.ArgumentParser(description="创建数据透视表")
    p.add_argument("input", help="输入 .xlsx"); p.add_argument("output", help="输出 .xlsx")
    p.add_argument("--rows", help="行字段"); p.add_argument("--cols", help="列字段")
    p.add_argument("--vals", help="值字段"); p.add_argument("--agg", default="sum", help="聚合方式: sum/avg/count/max/min")
    a = p.parse_args()

    wb = load_workbook(a.input, data_only=True); ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    data = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c+1).value for c in range(len(headers))}
        data.append(row)

    # 简单透视
    ri = headers.index(a.rows) if a.rows in headers else None
    ci = headers.index(a.cols) if a.cols in headers else None
    vi = headers.index(a.vals) if a.vals in headers else None

    if ri is None or not a.vals:
        print("需要有效的 --rows 和 --vals "); sys.exit(1)

    row_vals = sorted(set(d[a.rows] for d in data if d[a.rows] is not None))
    col_vals = sorted(set(d[a.cols] for d in data if a.cols and d[a.cols] is not None)) if ci is not None else ["总计"]

    pivot = {}
    for d in data:
        rk = d[a.rows]
        ck = d[a.cols] if ci is not None else "总计"
        val = d[a.vals]
        if rk is None or val is None: continue
        key = (rk, ck)
        if key not in pivot: pivot[key] = []
        if isinstance(val, (int, float)): pivot[key].append(val)

    # 输出
    out_wb = __import__('openpyxl').Workbook(); out_ws = out_wb.active
    out_ws.title = "透视表"
    out_ws.cell(1, 1, f"{a.rows} / {a.cols or '总计'}").font = __import__('openpyxl').styles.Font(bold=True, size=14)
    for j, cv in enumerate(col_vals):
        out_ws.cell(2, j+2, str(cv))

    agg_fn = {"sum": sum, "avg": lambda x: sum(x)/len(x), "count": len, "max": max, "min": min}.get(a.agg, sum)
    for i, rv in enumerate(row_vals):
        out_ws.cell(i+3, 1, str(rv))
        for j, cv in enumerate(col_vals):
            vals = pivot.get((rv, cv), [])
            out_ws.cell(i+3, j+2, agg_fn(vals) if vals else 0)

    out_wb.save(a.output)
    print(f"✅ 透视表已生成：{a.output}（{len(row_vals)} 行 × {len(col_vals)} 列）")

if __name__ == "__main__": main()
