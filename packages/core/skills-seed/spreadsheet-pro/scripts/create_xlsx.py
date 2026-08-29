#!/usr/bin/env python3
"""Otto Spreadsheet-Pro v6 — 视觉感知多布局引擎（含图​表）。python create_xlsx.py in.md out.xlsx"""
from __future__ import annotations
import re, sys
from pathlib import Path

if sys.platform == 'win32':
    try: sys.stdout.reconfigure(encoding='utf-8')
    except: pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

def _rgb(h): h=h.lstrip("#"); return tuple(int(h[i:i+2],16) for i in(0,2,4))
def _hex(r,g,b): return f"{max(0,min(255,int(r))):02X}{max(0,min(255,int(g))):02X}{max(0,min(255,int(b))):02X}"
def _readable_body(base): r,g,b=_rgb(base); lum=0.299*r+0.587*g+0.114*b; return _hex(r*0.35,g*0.35,b*0.35) if lum<80 else "2D2D2D"
def _readable_muted(base): r,g,b=_rgb(base); lum=0.299*r+0.587*g+0.114*b; return _hex(r*0.45+100,g*0.45+100,b*0.45+100) if lum<80 else "666666"

def resolve(meta):
    base=meta.get("base","0A1628"); accent=meta.get("accent","2D7DD2"); surface=meta.get("surface","F0F4F8")
    return {"theme":meta.get("theme",""),"atmo":meta.get("atmosphere",""),
        "base":base,"accent":accent,"surface":surface,
        "body":_readable_body(base),
        "muted":_readable_muted(base),
        "hdr_bg":"F0F2F5","hdr_text":base,"stripe":surface,
        "border":"D0D5DD","neg":"DC3545","pos":"28A745",
        "h_font":meta.get("heading_font","Microsoft YaHei"),
        "b_font":meta.get("body_font","Microsoft YaHei"),
        "t_sz":int(float(meta.get("title_size","12"))),
        "hd_sz":int(float(meta.get("header_size","10.5"))),
        "b_sz":int(float(meta.get("body_size","10")))}

LAYOUT_RE=re.compile(r'<!--\s*layout:\s*(\w[\w-]*)\s*-->')

def parse(text):
    meta={}; body=text.strip()
    if body.startswith("---"):
        parts=body.split("---",2)
        if len(parts)>=3:
            for line in parts[1].strip().split("\n"):
                if ":" in line and not line[0]=="#": k,_,v=line.partition(":"); meta[k.strip()]=v.strip().strip('"').strip("'")
            body=parts[2].strip()

    lines=body.split("\n"); i=0
    sheets=[]; cur={"name":"Sheet1","layout":"table","blocks":[],"rows":[]}
    def save():
        if cur["blocks"] or cur["rows"]: sheets.append(cur.copy())

    tbl=[]; in_t=False
    while i<len(lines):
        line=lines[i]
        if line.strip().startswith("|") and line.strip().endswith("|"):
            tbl.append(line); in_t=True; i+=1; continue
        elif in_t:
            h,r=_tbl(tbl)
            if h: cur["rows"]=[h]+r
            tbl=[]; in_t=False; continue
        if not line.strip(): i+=1; continue

        m=re.match(r"^(#{1,2})\s+(.+)$",line)
        if m:
            save(); txt=m.group(2).strip(); lm=LAYOUT_RE.search(txt)
            layout="table"
            if lm: layout=lm.group(1); txt=txt[:lm.start()].strip()
            cur={"name":txt[:31],"layout":layout,"blocks":[],"rows":[]}; i+=1; continue

        cur["blocks"].append({"t":"text","text":line.strip()})
        i+=1
    if tbl: h,r=_tbl(tbl);
    if h: cur["rows"]=[h]+r
    save()
    return meta,sheets

def _tbl(raw):
    if len(raw)<2: return [],[]
    h=[c.strip() for c in raw[0].strip("|").split("|")]
    rows=[]
    for line in raw[1:]:
        if re.match(r"^[\|\-\s:]+$",line.strip()): continue
        cells=[c.strip() for c in line.strip("|").split("|")]
        if cells: rows.append(cells)
    return h,rows


class R:
    def __init__(self,t,meta):
        self.t=t; self.m=meta; self.wb=Workbook(); self._first=True
        self._st()

    def _hex(self,k): return self.t[k]

    def _st(self):
        self._hf=Font(name=self.t["h_font"],size=self.t["hd_sz"],bold=True,color=self._hex("hdr_text"))
        self._hf_fill=PatternFill(start_color=self._hex("hdr_bg"),end_color=self._hex("hdr_bg"),fill_type="solid")
        self._bf=Font(name=self.t["b_font"],size=self.t["b_sz"],color=self._hex("body"))
        self._sf=PatternFill(start_color=self._hex("stripe"),end_color=self._hex("stripe"),fill_type="solid")
        self._tf=Font(name=self.t["h_font"],size=self.t["t_sz"],bold=True,color="FFFFFF")
        self._tf_fill=PatternFill(start_color=self._hex("base"),end_color=self._hex("base"),fill_type="solid")
        self._bdr=Border(left=Side(style="thin",color=self._hex("border")),
            right=Side(style="thin",color=self._hex("border")),
            top=Side(style="thin",color=self._hex("border")),
            bottom=Side(style="thin",color=self._hex("border")))
        self._ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
        self._lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
        self._pos_f=Font(name=self.t["b_font"],size=self.t["b_sz"],color=self._hex("pos"))
        self._neg_f=Font(name=self.t["b_font"],size=self.t["b_sz"],color=self._hex("neg"))

    def _ws(self,name):
        if self._first: ws=self.wb.active; ws.title=name; self._first=False
        else: ws=self.wb.create_sheet(title=name)
        ws.sheet_properties.tabColor=self._hex("accent")
        return ws

    def _col_w(self,ws,rows,cols):
        for j in range(cols):
            max_w=0
            for row in rows:
                if j<len(row):
                    w=sum(2.1 if ord(c)>127 else 1.1 for c in str(row[j]))
                    max_w=max(max_w,w)
            ws.column_dimensions[get_column_letter(j+1)].width=min(max_w+3,45)

    def _is_num(self,val):
        if not val or not isinstance(val,str): return False
        return bool(re.match(r"^[\-\+]?[\d,]+\.?\d*%?$",val.strip()))

    def _to_num(self,val):
        v=val.strip().replace(",","").replace("%","")
        try: return float(v)
        except: return None

    def build(self,sheets):
        title=self.m.get("title","")
        for si,sheet in enumerate(sheets):
            ws=self._ws(sheet["name"])
            layout=sheet.get("layout","table")

            if title and si==0:
                ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
                c=ws.cell(1,1,value=title); c.font=self._tf; c.fill=self._tf_fill; c.alignment=self._ctr
                ws.row_dimensions[1].height=26; sr=2
            else: sr=1

            r=sr; rows=sheet.get("rows",[])
            if not rows: continue

            # 文本说明
            for blk in sheet.get("blocks",[]):
                c=ws.cell(r,1,value=blk["text"][:80])
                c.font=Font(name=self.t["b_font"],size=self.t["b_sz"]-1,color=self._hex("muted"),italic=True); r+=1

            # dashboard 布局：大号KPI
            if layout=="dashboard" and len(rows)>=2:
                # 表头
                for j,h in enumerate(rows[0]):
                    c=ws.cell(sr+1,j+1,value=h); c.font=self._hf; c.fill=self._hf_fill; c.alignment=self._ctr; c.border=self._bdr
                ws.row_dimensions[sr+1].height=22
                # 数据（表头下方）
                for i,row in enumerate(rows[1:][:8]):
                    for j,val in enumerate(row):
                        c=ws.cell(sr+2+i,j+1,value=val)
                        c.font=Font(name=self.t["b_font"],size=self.t["b_sz"]+2 if j>0 else self.t["b_sz"],bold=(j>0),color=self._hex("base") if j>0 else self._hex("muted"))
                        if i%2==1: c.fill=self._sf
                        if j==0: c.alignment=self._lft
                        else: c.alignment=self._ctr
                        if j>0 and self._is_num(val):
                            num=self._to_num(val)
                            if num is not None:
                                if num<0: c.font=Font(name=self.t["b_font"],size=self.t["b_sz"]+2,bold=True,color=self._hex("neg"))
                                elif num>0: c.font=Font(name=self.t["b_font"],size=self.t["b_sz"]+2,bold=True,color=self._hex("pos"))
                self._col_w(ws,[rows[0]]+rows[1:],len(rows[0]))
                continue

            # 表头
            for j,h in enumerate(rows[0]):
                c=ws.cell(r,j+1,value=h); c.font=self._hf; c.fill=self._hf_fill; c.alignment=self._ctr; c.border=self._bdr
            ws.row_dimensions[r].height=20; r+=1; data_start=r

            for i,row in enumerate(rows[1:]):
                for j,val in enumerate(row):
                    c=ws.cell(r+i,j+1,value=val); c.font=self._bf; c.border=self._bdr; c.alignment=self._lft
                    if self._is_num(val):
                        c.alignment=self._ctr
                        num=self._to_num(val)
                        if num is not None:
                            if num<0: c.font=self._neg_f
                            elif num>0: c.font=self._pos_f
                    if i%2==1: c.fill=self._sf
                ws.row_dimensions[r+i].height=18

            self._col_w(ws,rows,len(rows[0]))
            if rows: ws.freeze_panes=ws.cell(sr,1)

            # chart 布局：自动生成柱状图
            if layout=="chart" and len(rows)>=2 and len(rows[0])>=2:
                try:
                    chart=BarChart(); chart.type="col"; chart.style=10
                    chart.title=sheet["name"]; chart.y_axis.title=None; chart.x_axis.title=None
                    cat=Reference(ws,min_col=1,min_row=data_start,max_row=data_start+len(rows)-2)
                    for cj in range(2,len(rows[0])+1):
                        val_ref=Reference(ws,min_col=cj,min_row=data_start-1,max_row=data_start+len(rows)-2)
                        chart.add_data(val_ref,titles_from_data=True)
                    chart.set_categories(cat)
                    chart.series[0].graphicalProperties.solidFill=self._hex("accent")
                    if len(chart.series)>1: chart.series[1].graphicalProperties.solidFill=self._hex("base")
                    chart.height=12; chart.width=20
                    ws.add_chart(chart,f"A{data_start+len(rows)+1}")
                except: pass

            # comparison 布局：条件格式数据条
            if layout=="comparison" and len(rows)>=2 and len(rows[0])>=2:
                try:
                    for cj in range(2,len(rows[0])+1):
                        col_letter=get_column_letter(cj)
                        rng=f"{col_letter}{data_start}:{col_letter}{data_start+len(rows)-2}"
                        ws.conditional_formatting.add(rng,
                            DataBarRule(start_type="min",end_type="max",color=self._hex("accent"),showValue=True))
                except: pass

    def save(self,p): self.wb.save(p)


def main():
    import argparse
    p=argparse.ArgumentParser(description="Otto Spreadsheet-Pro v6")
    p.add_argument("input"); p.add_argument("output")
    a=p.parse_args()
    ip=Path(a.input)
    if not ip.exists(): print(f"找不到 {a.input}"); sys.exit(1)
    meta,sheets=parse(ip.read_text(encoding="utf-8"))
    t=resolve(meta)
    if "title" not in meta: meta["title"]=ip.stem
    g=R(t,meta); g.build(sheets)
    op=Path(a.output); op.parent.mkdir(parents=True,exist_ok=True)
    g.save(str(op))
    lnames={"table":"标准","dashboard":"看板","chart":"图表","comparison":"对比"}
    layout_summary=", ".join(f"{s['name'][:8]}→{lnames.get(s.get('layout','table'),s.get('layout','?'))}" for s in sheets)
    print(f"✅ {op.name}  {op.stat().st_size/1024:.0f}KB  {len(sheets)}sheets")
    print(f"   布局: {layout_summary}")

if __name__=="__main__": main()
